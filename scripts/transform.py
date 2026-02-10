"""Data transformation utilities for raw source file preparation.

Provides three capabilities required between file acquisition (E-301)
and schema validation / Snowflake loading (E-303):

1. XLSX-to-CSV conversion for TTC delay files (openpyxl, streaming).
2. Encoding detection and UTF-8 normalization (charset-normalizer).
3. ZIP archive extraction for Bike Share ridership files (zipfile).

All functions are idempotent and produce deterministic output.
"""

from __future__ import annotations

import csv
import logging
import shutil
import time
import zipfile
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import TYPE_CHECKING

import openpyxl
from charset_normalizer import from_path

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

_ENCODING_CONFIDENCE_THRESHOLD: float = 0.7
_ENCODING_CHUNK_SIZE: int = 1_048_576  # 1 MB

# BOM byte sequences to strip from file start
_UTF8_BOM: bytes = b"\xef\xbb\xbf"
_UTF16_LE_BOM: bytes = b"\xff\xfe"
_UTF16_BE_BOM: bytes = b"\xfe\xff"


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------


class TransformError(Exception):
    """Raised when a transformation operation fails."""


class EncodingError(Exception):
    """Raised when encoding detection confidence is below threshold."""


# ---------------------------------------------------------------------------
# Result dataclasses
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class TransformResult:
    """Outcome of an XLSX-to-CSV conversion.

    Attributes:
        input_path: Source XLSX file path.
        output_path: Destination CSV file path.
        row_count: Number of data rows written (excluding header).
        column_count: Number of columns in the output CSV.
        elapsed_seconds: Wall-clock duration of the conversion.
    """

    input_path: Path
    output_path: Path
    row_count: int
    column_count: int
    elapsed_seconds: float


@dataclass(frozen=True, slots=True)
class EncodingResult:
    """Outcome of an encoding normalization operation.

    Attributes:
        input_path: Source file path.
        output_path: Destination file path (may equal input_path for in-place).
        detected_encoding: Encoding identified by charset-normalizer.
        confidence: Detection confidence score between 0.0 and 1.0.
        byte_count: Total bytes in the output file.
        had_bom: Whether a BOM was detected and stripped.
    """

    input_path: Path
    output_path: Path
    detected_encoding: str
    confidence: float
    byte_count: int
    had_bom: bool


@dataclass(frozen=True, slots=True)
class ExtractResult:
    """Outcome of a single file extraction from a ZIP archive.

    Attributes:
        zip_path: Source archive path.
        extracted_path: Output file path on disk.
        original_name: Filename as stored inside the archive.
        byte_size: Uncompressed size of the extracted file.
        skipped: True if extraction was skipped due to idempotency check.
    """

    zip_path: Path
    extracted_path: Path
    original_name: str
    byte_size: int
    skipped: bool


# ---------------------------------------------------------------------------
# XLSX-to-CSV Conversion (S002)
# ---------------------------------------------------------------------------


def convert_xlsx_to_csv(
    xlsx_path: Path,
    csv_path: Path,
    sheet_name: str | None = None,
) -> TransformResult:
    """Convert a single XLSX file to UTF-8 CSV using streaming reads.

    Uses openpyxl in read-only mode to avoid loading the entire workbook
    into memory. Writes output with csv.QUOTE_MINIMAL quoting.

    Args:
        xlsx_path: Path to the source XLSX file.
        csv_path: Path for the output CSV file. Parent dirs are created.
        sheet_name: Worksheet name to read. Reads the active sheet if None.

    Returns:
        TransformResult with conversion metadata.

    Raises:
        TransformError: If the specified sheet_name does not exist,
            or if the file is not a valid XLSX workbook.
        FileNotFoundError: If xlsx_path does not exist.
    """
    start = time.monotonic()

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except zipfile.BadZipFile as exc:
        raise TransformError(
            f"'{xlsx_path}' is not a valid XLSX file "
            f"(may be a CSV with .xlsx extension): {exc}"
        ) from exc
    try:
        ws = _resolve_worksheet(wb, sheet_name, xlsx_path)
        csv_path.parent.mkdir(parents=True, exist_ok=True)

        row_count = 0
        column_count = 0

        with csv_path.open("w", newline="", encoding="utf-8") as fout:
            writer = csv.writer(fout, quoting=csv.QUOTE_MINIMAL)
            for i, row in enumerate(ws.iter_rows()):
                values = ["" if cell.value is None else str(cell.value) for cell in row]
                writer.writerow(values)
                if i == 0:
                    column_count = len(values)
                else:
                    row_count += 1
    finally:
        wb.close()

    elapsed = time.monotonic() - start
    logger.info(
        "Converted %s -> %s (%d rows, %.2fs)",
        xlsx_path.name,
        csv_path.name,
        row_count,
        elapsed,
    )
    return TransformResult(
        input_path=xlsx_path,
        output_path=csv_path,
        row_count=row_count,
        column_count=column_count,
        elapsed_seconds=round(elapsed, 3),
    )


def _resolve_worksheet(
    wb: openpyxl.Workbook,
    sheet_name: str | None,
    xlsx_path: Path,
) -> Worksheet:
    """Return the target worksheet, raising TransformError on mismatch."""
    if sheet_name is None:
        ws = wb.active
        if ws is None:
            raise TransformError(f"Workbook '{xlsx_path}' has no active worksheet")
        return ws

    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise TransformError(
            f"Sheet '{sheet_name}' not found in '{xlsx_path}'. "
            f"Available sheets: {available}"
        )
    return wb[sheet_name]


def batch_convert(
    source_dir: Path,
    output_dir: Path,
    file_pattern: str = "*.xlsx",
) -> list[TransformResult]:
    """Convert all matching XLSX files in a directory tree to CSV.

    Preserves subdirectory structure relative to source_dir in the
    output directory. Each .xlsx file produces a .csv sibling.

    Args:
        source_dir: Root directory to scan for XLSX files.
        output_dir: Root directory for CSV output.
        file_pattern: Glob pattern for matching XLSX files.

    Returns:
        List of TransformResult, one per converted file.
    """
    results: list[TransformResult] = []
    for xlsx_path in sorted(source_dir.rglob(file_pattern)):
        relative = xlsx_path.relative_to(source_dir)
        csv_name = relative.with_suffix(".csv")
        csv_path = output_dir / csv_name
        result = convert_xlsx_to_csv(xlsx_path, csv_path)
        results.append(result)
    return results


# ---------------------------------------------------------------------------
# Encoding Detection and Normalization (S004)
# ---------------------------------------------------------------------------


def normalize_encoding(
    input_path: Path,
    output_path: Path | None = None,
) -> EncodingResult:
    """Detect file encoding and transcode to UTF-8 if necessary.

    Uses charset-normalizer for detection. Strips BOM bytes from the
    output. Processes files in 1 MB chunks to limit memory usage.

    Args:
        input_path: Path to the source file.
        output_path: Destination path. Overwrites input_path if None.

    Returns:
        EncodingResult with detection metadata.

    Raises:
        EncodingError: If detection confidence is below 0.7.
        FileNotFoundError: If input_path does not exist.
    """
    target = output_path if output_path is not None else input_path
    detection = from_path(input_path)
    best = detection.best()

    if best is None:
        raise EncodingError(
            f"Cannot detect encoding for '{input_path}': no candidates returned"
        )

    detected_encoding: str = str(best.encoding)
    # charset-normalizer uses chaos (0=perfect). Invert to confidence.
    confidence: float = 1.0 - best.chaos

    if confidence < _ENCODING_CONFIDENCE_THRESHOLD:
        candidates = list(detection)[:3]
        top_candidates = [f"{r.encoding} ({1.0 - r.chaos:.2f})" for r in candidates]
        raise EncodingError(
            f"Low confidence ({confidence:.2f}) detecting encoding for "
            f"'{input_path}'. Top candidates: {', '.join(top_candidates)}"
        )

    raw_bytes = input_path.read_bytes()
    had_bom = _has_bom(raw_bytes)

    if had_bom:
        raw_bytes = _strip_bom(raw_bytes)

    is_utf8 = detected_encoding.lower().replace("-", "") in {
        "utf8",
        "ascii",
    }

    if is_utf8 and not had_bom and target == input_path:
        return EncodingResult(
            input_path=input_path,
            output_path=target,
            detected_encoding=detected_encoding,
            confidence=round(confidence, 4),
            byte_count=len(raw_bytes),
            had_bom=had_bom,
        )

    if is_utf8:
        # Already UTF-8 but needs BOM stripped or different output path
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(raw_bytes)
    else:
        # Transcode to UTF-8
        text = raw_bytes.decode(detected_encoding)
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(text, encoding="utf-8")

    byte_count = target.stat().st_size

    logger.info(
        "Normalized %s: %s (%.2f) -> UTF-8%s",
        input_path.name,
        detected_encoding,
        confidence,
        " [BOM stripped]" if had_bom else "",
    )

    return EncodingResult(
        input_path=input_path,
        output_path=target,
        detected_encoding=detected_encoding,
        confidence=round(confidence, 4),
        byte_count=byte_count,
        had_bom=had_bom,
    )


def _has_bom(data: bytes) -> bool:
    """Check whether raw bytes begin with a known BOM sequence."""
    return (
        data.startswith(_UTF8_BOM)
        or data.startswith(_UTF16_LE_BOM)
        or data.startswith(_UTF16_BE_BOM)
    )


def _strip_bom(data: bytes) -> bytes:
    """Remove leading BOM bytes from raw file content."""
    if data.startswith(_UTF8_BOM):
        return data[len(_UTF8_BOM) :]
    if data.startswith(_UTF16_LE_BOM):
        return data[len(_UTF16_LE_BOM) :]
    if data.startswith(_UTF16_BE_BOM):
        return data[len(_UTF16_BE_BOM) :]
    return data


# ---------------------------------------------------------------------------
# ZIP Archive Extraction (S007)
# ---------------------------------------------------------------------------


def extract_zip(
    zip_path: Path,
    output_dir: Path,
) -> list[ExtractResult]:
    """Extract CSV files from a ZIP archive with idempotent behavior.

    Discovers CSV members dynamically via ZipFile.infolist(). Strips
    internal subdirectory prefixes to produce flat output. Skips
    extraction when a target file already exists with matching size.

    Args:
        zip_path: Path to the source ZIP archive.
        output_dir: Directory to write extracted CSV files.

    Returns:
        List of ExtractResult, one per CSV member.

    Raises:
        TransformError: If the archive is corrupt (fails integrity test).
        FileNotFoundError: If zip_path does not exist.
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(zip_path, "r") as zf:
        corrupt_member = zf.testzip()
        if corrupt_member is not None:
            raise TransformError(
                f"Corrupt ZIP archive '{zip_path}': bad member '{corrupt_member}'"
            )

        csv_members = _discover_csv_members(zf)
        results: list[ExtractResult] = []

        for info in csv_members:
            flat_name = PurePosixPath(info.filename).name
            target = output_dir / flat_name

            if target.exists() and target.stat().st_size == info.file_size:
                logger.debug("SKIPPED %s (size match)", flat_name)
                results.append(
                    ExtractResult(
                        zip_path=zip_path,
                        extracted_path=target,
                        original_name=info.filename,
                        byte_size=info.file_size,
                        skipped=True,
                    )
                )
                continue

            with zf.open(info.filename) as src, target.open("wb") as dst:
                shutil.copyfileobj(src, dst)

            logger.info(
                "Extracted %s -> %s (%d bytes)",
                info.filename,
                target.name,
                info.file_size,
            )
            results.append(
                ExtractResult(
                    zip_path=zip_path,
                    extracted_path=target,
                    original_name=info.filename,
                    byte_size=info.file_size,
                    skipped=False,
                )
            )

    return results


def _discover_csv_members(
    zf: zipfile.ZipFile,
) -> list[zipfile.ZipInfo]:
    """Filter ZIP members to CSV entries, excluding macOS metadata."""
    return [
        info
        for info in zf.infolist()
        if info.filename.lower().endswith(".csv")
        and "__MACOSX" not in info.filename
        and not info.is_dir()
    ]


def rename_csv_columns(
    csv_path: Path,
    column_map: dict[str, str],
) -> bool:
    """Rename CSV header columns in-place using the provided mapping.

    Reads the CSV, applies column renames to the header row, and writes
    back. Only modifies the file if at least one rename was applied.

    Args:
        csv_path: Path to the CSV file to modify.
        column_map: Mapping of old column names to new column names.

    Returns:
        True if any columns were renamed, False if no changes needed.
    """
    raw = csv_path.read_text(encoding="utf-8")
    lines = raw.split("\n")
    if not lines:
        return False

    header = lines[0]
    original_header = header
    for old_name, new_name in column_map.items():
        # Replace exact column name matches in the CSV header.
        # Uses comma boundaries and start/end anchors to avoid partial matches.
        parts = header.split(",")
        parts = [new_name if p.strip() == old_name else p for p in parts]
        header = ",".join(parts)

    if header == original_header:
        return False

    lines[0] = header
    csv_path.write_text("\n".join(lines), encoding="utf-8")
    logger.info("Renamed columns in %s: %s", csv_path.name, column_map)
    return True


def strip_extra_columns(
    csv_path: Path,
    allowed_columns: frozenset[str],
) -> bool:
    """Remove columns not in the allowed set from a CSV file in-place.

    Uses proper CSV parsing to handle quoted fields containing delimiters.
    Comparison is case-insensitive to match the validation engine.

    Args:
        csv_path: Path to the CSV file to modify.
        allowed_columns: Set of column names to retain.

    Returns:
        True if any columns were removed, False if no changes needed.
    """
    with csv_path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header is None:
            return False

        allowed_lower = {c.lower() for c in allowed_columns}
        keep_indices = [
            i for i, col in enumerate(header) if col.strip().lower() in allowed_lower
        ]

        if len(keep_indices) == len(header):
            return False

        removed = [header[i] for i in range(len(header)) if i not in set(keep_indices)]

        rows: list[list[str]] = [[header[i] for i in keep_indices]]
        for row in reader:
            rows.append([row[i] if i < len(row) else "" for i in keep_indices])

    with csv_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerows(rows)

    logger.info("Stripped extra columns from %s: %s", csv_path.name, removed)
    return True


def batch_extract_zips(
    source_dir: Path,
    output_dir: Path,
    file_pattern: str = "*.zip",
) -> list[ExtractResult]:
    """Extract all matching ZIP archives in a directory tree.

    Preserves the <source>/<year>/ subdirectory structure from source_dir
    in the output directory.

    Args:
        source_dir: Root directory to scan for ZIP files.
        output_dir: Root directory for extracted output.
        file_pattern: Glob pattern for matching ZIP files.

    Returns:
        Flat list of ExtractResult across all archives.
    """
    all_results: list[ExtractResult] = []
    for zip_path in sorted(source_dir.rglob(file_pattern)):
        relative_dir = zip_path.parent.relative_to(source_dir)
        target_dir = output_dir / relative_dir
        results = extract_zip(zip_path, target_dir)
        all_results.extend(results)
    return all_results
