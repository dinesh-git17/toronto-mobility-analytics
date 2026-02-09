"""Shared pytest fixtures for validation and transformation tests.

Generates test fixtures programmatically to avoid committing binary
files. XLSX fixtures use openpyxl; CSV fixtures use csv.writer;
encoding fixtures use explicit byte encoding.
"""

from __future__ import annotations

import csv
import zipfile
from typing import TYPE_CHECKING

import openpyxl
import pytest

if TYPE_CHECKING:
    from pathlib import Path

# ---------------------------------------------------------------------------
# TTC Subway delay data (matches DESIGN-DOC Section 4.3.1)
# ---------------------------------------------------------------------------

SUBWAY_HEADERS: list[str] = [
    "Date",
    "Time",
    "Day",
    "Station",
    "Code",
    "Min Delay",
    "Min Gap",
    "Bound",
    "Line",
    "Vehicle",
]

SUBWAY_ROWS: list[list[str]] = [
    # Date, Time, Day, Station, Code, MinDelay, MinGap, Bound, Line, Vehicle
    [
        "2023-01-01",
        "08:30",
        "Sunday",
        "BLOOR STATION",
        "MUPAA",
        "5",
        "10",
        "N",
        "YU",
        "5931",
    ],
    [
        "2023-01-01",
        "09:15",
        "Sunday",
        "MUSEUM STATION",
        "PUOPO",
        "3",
        "8",
        "S",
        "YU",
        "5904",
    ],
    [
        "2023-01-02",
        "07:00",
        "Monday",
        "KENNEDY STATION",
        "MUNCA",
        "12",
        "",
        "E",
        "BD",
        "6021",
    ],
    ["2023-01-02", "12:30", "Monday", "FINCH STATION", "", "2", "5", "", "YU", "5843"],
    [
        "2023-01-03",
        "16:45",
        "Tuesday",
        "ST GEORGE STATION",
        "MUPAA",
        "8",
        "15",
        "N",
        "YU",
        "5901",
    ],
    [
        "2023-01-04",
        "06:00",
        "Wednesday",
        "UNION STATION",
        "SUDP",
        "20",
        "25",
        "S",
        "YU",
        "5876",
    ],
    [
        "2023-01-04",
        "18:00",
        "Wednesday",
        "KIPLING STATION",
        "MUNCA",
        "4",
        "7",
        "W",
        "BD",
        "6012",
    ],
    [
        "2023-01-05",
        "08:15",
        "Thursday",
        "EGLINTON STATION",
        "TUSC",
        "6",
        "9",
        "N",
        "YU",
        "5899",
    ],
    [
        "2023-01-06",
        "09:30",
        "Friday",
        "SHEPPARD STATION",
        "MUPAA",
        "3",
        "6",
        "N",
        "SHP",
        "6101",
    ],
    [
        "2023-01-07",
        "14:00",
        "Saturday",
        "DUNDAS STATION",
        "PUOPO",
        "7",
        "12",
        "S",
        "YU",
        "5888",
    ],
]


@pytest.fixture()
def valid_subway_csv(tmp_path: Path) -> Path:
    """Create a valid 10-row CSV matching the TTC subway contract."""
    csv_path = tmp_path / "valid_subway_delays.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(SUBWAY_HEADERS)
        writer.writerows(SUBWAY_ROWS)
    return csv_path


@pytest.fixture()
def invalid_missing_column_csv(tmp_path: Path) -> Path:
    """Create a subway CSV with the 'Station' column removed."""
    headers = [h for h in SUBWAY_HEADERS if h != "Station"]
    csv_path = tmp_path / "invalid_missing_column.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in SUBWAY_ROWS:
            # Remove index 3 (Station)
            writer.writerow(row[:3] + row[4:])
    return csv_path


@pytest.fixture()
def invalid_type_mismatch_csv(tmp_path: Path) -> Path:
    """Create a subway CSV where Min Delay contains non-integer values."""
    csv_path = tmp_path / "invalid_type_mismatch.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(SUBWAY_HEADERS)
        # First row: valid
        writer.writerow(SUBWAY_ROWS[0])
        # Second row: "abc" in Min Delay (index 5)
        bad_row = list(SUBWAY_ROWS[1])
        bad_row[5] = "abc"
        writer.writerow(bad_row)
    return csv_path


@pytest.fixture()
def extra_column_csv(tmp_path: Path) -> Path:
    """Create a subway CSV with an extra 'Notes' column."""
    headers = [*SUBWAY_HEADERS, "Notes"]
    csv_path = tmp_path / "extra_column.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in SUBWAY_ROWS:
            writer.writerow([*row, "extra data"])
    return csv_path


@pytest.fixture()
def case_mismatch_csv(tmp_path: Path) -> Path:
    """Create a subway CSV where headers use different casing."""
    headers = [h.upper() for h in SUBWAY_HEADERS]
    csv_path = tmp_path / "case_mismatch.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(SUBWAY_ROWS)
    return csv_path


@pytest.fixture()
def nullable_empty_csv(tmp_path: Path) -> Path:
    """Create a subway CSV where nullable columns have empty values."""
    csv_path = tmp_path / "nullable_empty.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(SUBWAY_HEADERS)
        for row in SUBWAY_ROWS:
            # Code (idx 4), Min Gap (idx 7), Bound (idx 7) are nullable
            modified = list(row)
            modified[4] = ""  # Code
            modified[7] = ""  # Bound
            writer.writerow(modified)
    return csv_path


# ---------------------------------------------------------------------------
# XLSX fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def valid_xlsx(tmp_path: Path) -> Path:
    """Create a small XLSX file with 10 rows of subway delay data."""
    xlsx_path = tmp_path / "valid_delays.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(SUBWAY_HEADERS)
    for row in SUBWAY_ROWS:
        ws.append(row)
    wb.save(xlsx_path)
    wb.close()
    return xlsx_path


@pytest.fixture()
def multi_sheet_xlsx(tmp_path: Path) -> Path:
    """Create an XLSX file with two named worksheets."""
    xlsx_path = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "delays"
    ws1.append(SUBWAY_HEADERS)
    for row in SUBWAY_ROWS:
        ws1.append(row)
    ws2 = wb.create_sheet("metadata")
    ws2.append(["key", "value"])
    ws2.append(["version", "2023"])
    wb.save(xlsx_path)
    wb.close()
    return xlsx_path


# ---------------------------------------------------------------------------
# Encoding fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def windows_1252_csv(tmp_path: Path) -> Path:
    """Create a CSV file encoded in Windows-1252 with special characters."""
    csv_path = tmp_path / "windows_1252.csv"
    content = "Name,Value\nCafé Résumé,123\nNaïve,456\n"
    csv_path.write_bytes(content.encode("windows-1252"))
    return csv_path


@pytest.fixture()
def utf8_bom_csv(tmp_path: Path) -> Path:
    """Create a UTF-8 CSV with a BOM prefix."""
    csv_path = tmp_path / "utf8_bom.csv"
    bom = b"\xef\xbb\xbf"
    content = "Name,Value\nTest,123\n"
    csv_path.write_bytes(bom + content.encode("utf-8"))
    return csv_path


# ---------------------------------------------------------------------------
# ZIP fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def valid_zip(tmp_path: Path) -> Path:
    """Create a ZIP archive containing two CSV files with subdirectory prefix."""
    zip_path = tmp_path / "bikeshare-ridership-2023.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        csv1 = "Trip Id,Trip  Duration\n1,600\n2,900\n"
        zf.writestr("bikeshare-ridership-2023/Bike share ridership 2023-01.csv", csv1)
        csv2 = "Trip Id,Trip  Duration\n3,1200\n4,300\n"
        zf.writestr("bikeshare-ridership-2023/Bike share ridership 2023-02.csv", csv2)
    return zip_path


@pytest.fixture()
def zip_with_macosx(tmp_path: Path) -> Path:
    """Create a ZIP with __MACOSX metadata that should be skipped."""
    zip_path = tmp_path / "with_macosx.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr("data/file.csv", "col1,col2\na,b\n")
        zf.writestr("__MACOSX/._file.csv", "metadata")
    return zip_path


@pytest.fixture()
def corrupt_zip(tmp_path: Path) -> Path:
    """Create a corrupt ZIP file with invalid CRC for a member."""
    zip_path = tmp_path / "corrupt.zip"
    # Write valid ZIP with enough data to reliably corrupt
    content = "a,b\n" + "1,2\n" * 500
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("data.csv", content)
    raw = bytearray(zip_path.read_bytes())
    # Corrupt the compressed data region (after local file header, ~40 bytes in)
    for i in range(50, min(100, len(raw))):
        raw[i] ^= 0xFF
    zip_path.write_bytes(bytes(raw))
    return zip_path
